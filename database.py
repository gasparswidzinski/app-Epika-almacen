# database.py
import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from datetime import datetime
import os
from pathlib import Path
import os as _os, hashlib as _hashlib, binascii as _binascii




# --- Rutas robustas: usar ProgramData para que siempre sea visible ---
import os, sys, sqlite3
from datetime import datetime

APP_NAME = "GestorDeStock"

def _get_progdata() -> str:
    base = os.environ.get("PROGRAMDATA") or os.path.join(os.path.expanduser("~"), "AppData", "Local")
    target = os.path.join(base, APP_NAME)
    os.makedirs(target, exist_ok=True)
    return target

def _scan_msstore_localcache_candidates() -> list:
    """Origen típico de Python MS Store: ...\Local\Packages\PythonSoftwareFoundation...\LocalCache\Roaming\GestorDeStock"""
    cands = []
    local = os.environ.get("LOCALAPPDATA", "")
    pkgs = os.path.join(local, "Packages")
    try:
        for name in os.listdir(pkgs):
            if name.startswith("PythonSoftwareFoundation.Python"):
                cands.append(os.path.join(pkgs, name, "LocalCache", "Roaming", APP_NAME))
    except Exception:
        pass
    return cands

def _migrar_a(dest_dir: str):
    """Si hay DB en orígenes conocidos, copiarla a dest_dir/almacen.db (una sola vez)."""
    import shutil
    new_db = os.path.join(dest_dir, "almacen.db")
    if os.path.exists(new_db):
        return  # ya hay DB en destino

    # 1) Desde MS Store LocalCache\Roaming
    for cand in _scan_msstore_localcache_candidates():
        old_db = os.path.join(cand, "almacen.db")
        if os.path.exists(old_db):
            try:
                shutil.copy(old_db, new_db)
                return
            except Exception as e:
                print("⚠️ No se pudo migrar desde LocalCache:", e)

    # 2) Desde Roaming real
    roaming = os.path.join(os.environ.get("USERPROFILE", ""), "AppData", "Roaming", APP_NAME)
    old_db = os.path.join(roaming, "almacen.db")
    if os.path.exists(old_db):
        try:
            shutil.copy(old_db, new_db)
            return
        except Exception as e:
            print("⚠️ No se pudo migrar desde Roaming:", e)

    # 3) Desde la carpeta del ejecutable (instalaciones viejas)
    try:
        exe_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
        legacy = os.path.join(exe_dir, "almacen.db")
        if os.path.exists(legacy):
            bkp = os.path.join(exe_dir, f"almacen_legacy_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
            shutil.copy(legacy, bkp)          # respaldo junto al exe
            shutil.copy(legacy, new_db)       # migración efectiva
    except Exception as e:
        print("⚠️ No se pudo migrar desde la carpeta del ejecutable:", e)

# Directorio de datos definitivo (ProgramData)
DATA_DIR = _get_progdata()
_migrar_a(DATA_DIR)

DB_PATH = os.path.join(DATA_DIR, "almacen.db")

def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def _migrar_db_si_corresponde():
    """
    Si existe 'almacen.db' junto al código/ejecutable (instalaciones viejas) y NO existe en DATA_DIR,
    copiamos la DB a la nueva ubicación y dejamos un backup junto al exe.
    """
    try:
        exe_dir = os.path.dirname(os.path.abspath(__file__))
        vieja = os.path.join(exe_dir, "almacen.db")
        if os.path.exists(vieja) and not os.path.exists(DB_PATH):
            import shutil
            bkp = os.path.join(exe_dir, f"almacen_legacy_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
            shutil.copy(vieja, bkp)     # respaldo por las dudas
            shutil.copy(vieja, DB_PATH) # migración efectiva
    except Exception as e:
        print("⚠️ No se pudo migrar DB:", e)


# -----------------------------
# Inicialización / migración
# -----------------------------
def inicializar_db():
    _migrar_db_si_corresponde()
    conn = get_connection()
    cur = conn.cursor()

    # Tabla sectores (si no existe)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS sectores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT UNIQUE NOT NULL,
        margen REAL NOT NULL
    )
    """)

    # Tabla productos
    cur.execute("""
    CREATE TABLE IF NOT EXISTS productos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo TEXT UNIQUE NOT NULL,
        nombre TEXT NOT NULL,
        cantidad INTEGER NOT NULL DEFAULT 0,
        costo REAL DEFAULT 0,
        sector_id INTEGER,
        precio REAL DEFAULT 0,
        codigo_barras TEXT,
        movimientos INTEGER DEFAULT 0,
        FOREIGN KEY(sector_id) REFERENCES sectores(id)
    )
    """)

    # Tabla movimientos (historial)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS movimientos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        producto_id INTEGER,
        tipo TEXT NOT NULL,
        cambio INTEGER NOT NULL,
        precio_unitario REAL NOT NULL,
        fecha TEXT NOT NULL,
        detalles TEXT,
        FOREIGN KEY(producto_id) REFERENCES productos(id)
    )
    """)

    # Tabla ventas (mantenemos la columna 'cliente' por compatibilidad)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS ventas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT NOT NULL,
        tipo_pago TEXT NOT NULL,
        estado TEXT NOT NULL,
        total REAL NOT NULL,
        efectivo_recibido REAL,
        vuelto REAL,
        cliente TEXT
    )
    """)

    # Si no existe la columna cliente_id la agregamos (migración)
    cur.execute("PRAGMA table_info(ventas)")
    cols = [r[1] for r in cur.fetchall()]
    if "cliente_id" not in cols:
        try:
            cur.execute("ALTER TABLE ventas ADD COLUMN cliente_id INTEGER")
        except Exception:
            pass

    # Tabla items de venta
    cur.execute("""
    CREATE TABLE IF NOT EXISTS venta_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        venta_id INTEGER NOT NULL,
        producto_id INTEGER NOT NULL,
        cantidad INTEGER NOT NULL,
        precio_unitario REAL NOT NULL,
        subtotal REAL NOT NULL,
        FOREIGN KEY(venta_id) REFERENCES ventas(id),
        FOREIGN KEY(producto_id) REFERENCES productos(id)
    )
    """)

    # Tabla clientes
    cur.execute("""
    CREATE TABLE IF NOT EXISTS clientes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        telefono TEXT,
        direccion TEXT,
        notas TEXT
    )
    """)

    # Tabla cobros (cuando se cobran ventas pendientes)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS cobros (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        venta_id INTEGER,
        cliente_id INTEGER,
        fecha TEXT,
        monto REAL,
        tipo_pago TEXT,
        FOREIGN KEY(venta_id) REFERENCES ventas(id),
        FOREIGN KEY(cliente_id) REFERENCES clientes(id)
    )
    """)

    # Semillas sectores (si no existen)
    default = [
        ("Fiambreria", 0.60),
        ("Panaderia", 0.40),
        ("Lacteos", 0.35),
        ("Almacen", 0.30)
    ]
    for nombre, margen in default:
        try:
            cur.execute("INSERT INTO sectores (nombre, margen) VALUES (?, ?)", (nombre, margen))
        except sqlite3.IntegrityError:
            pass

    # Tabla gastos (almacén y personal)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS gastos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT NOT NULL,
        categoria TEXT NOT NULL,
        monto REAL NOT NULL,
        detalle TEXT,
        tipo TEXT NOT NULL CHECK(tipo IN ('almacen','personal'))
    )
    """)
    
    # --- Nueva tabla para carrito temporal ---
    cur.execute("""
    CREATE TABLE IF NOT EXISTS carrito_temporal (
        producto_id INTEGER,
        codigo TEXT,
        nombre TEXT,
        cantidad REAL,
        precio_unitario REAL
    )
    """)
    
        # Tabla categorías de gastos
    cur.execute("""
    CREATE TABLE IF NOT EXISTS categorias_gasto (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT UNIQUE NOT NULL,
        tipo TEXT NOT NULL CHECK(tipo IN ('almacen','personal'))
    )
    """)

    # Semillas de categorías de gastos (solo si no existen)
    default_cats_almacen = [
        "Proveedores", "Sueldos", "Alquiler", "Luz",
        "Impuestos", "Contador", "Agua", "Otros"
    ]
    default_cats_personal = [
        "Alquiler Vivienda", "Comida", "Transporte",
        "Educación", "Salud", "Entretenimiento", "Otros"
    ]

    for nombre in default_cats_almacen:
        try:
            cur.execute("INSERT INTO categorias_gasto (nombre, tipo) VALUES (?, ?)", (nombre, "almacen"))
        except sqlite3.IntegrityError:
            pass

    for nombre in default_cats_personal:
        try:
            cur.execute("INSERT INTO categorias_gasto (nombre, tipo) VALUES (?, ?)", (nombre, "personal"))
        except sqlite3.IntegrityError:
            pass
        
     # Tabla usuarios
    cur.execute("""
    CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        rol TEXT NOT NULL CHECK(rol IN ('admin','user','developer'))

    )
    """)

    # --- Asegurar usuarios por defecto (si faltan) ---
    try:
        cur.execute("SELECT 1 FROM usuarios WHERE usuario='admin'")
        if not cur.fetchone():
            cur.execute("INSERT INTO usuarios (usuario,password,rol) VALUES (?,?,?)",
                        ('admin', _hash_password('admin'), 'admin'))
        cur.execute("SELECT 1 FROM usuarios WHERE usuario='developer'")
        if not cur.fetchone():
            cur.execute("INSERT INTO usuarios (usuario,password,rol) VALUES (?,?,?)",
                        ('developer', _hash_password('developer'), 'developer'))
    except Exception as e:
        print("⚠️ No se pudo asegurar usuarios por defecto:", e)

        
    # --- Índice único condicional para código de barras (evita duplicados no nulos) ---
    try:
        cur.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_prod_barcode
            ON productos(codigo_barras)
            WHERE codigo_barras IS NOT NULL
        """)
    except Exception:
        # Si existen duplicados hoy, el índice no se creará.
        # La app sigue funcionando; al depurar duplicados, se creará en el próximo arranque.
        pass
    
    conn.commit()
    conn.close()

# Alias
init_db = inicializar_db

# -----------------------------
# SECTORES
# -----------------------------
def obtener_sectores():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, nombre, margen FROM sectores ORDER BY nombre")
    data = cur.fetchall()
    conn.close()
    return data

def agregar_sector(nombre, margen):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("INSERT INTO sectores (nombre, margen) VALUES (?, ?)", (nombre, margen))
    conn.commit()
    conn.close()

def editar_sector(id_sector, nombre, margen):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE sectores SET nombre=?, margen=? WHERE id=?", (nombre, margen, id_sector))
    conn.commit()
    conn.close()

def eliminar_sector(id_sector):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM sectores WHERE id=?", (id_sector,))
    conn.commit()
    conn.close()

def obtener_margen_sector(id_sector):
    if id_sector is None:
        return 0.0
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT margen FROM sectores WHERE id=?", (id_sector,))
    row = cur.fetchone()
    conn.close()
    return row[0] if row else 0.0

# -----------------------------
# MOVIMIENTOS
# -----------------------------
def agregar_movimiento(producto_id, tipo, cambio, precio_unitario, detalles=""):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO movimientos (producto_id, tipo, cambio, precio_unitario, fecha, detalles) VALUES (?, ?, ?, ?, ?, ?)",
        (producto_id, tipo, cambio, precio_unitario, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), detalles)
    )
    conn.commit()
    conn.close()

def obtener_movimientos(limit=100):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT m.id, p.nombre, m.tipo, m.cambio, m.precio_unitario, m.fecha, m.detalles
        FROM movimientos m
        LEFT JOIN productos p ON p.id = m.producto_id
        ORDER BY m.id DESC
        LIMIT ?
    """, (limit,))
    data = cur.fetchall()
    conn.close()
    return data

# -----------------------------
# PRODUCTOS
# -----------------------------
def agregar_producto(codigo, nombre, cantidad, costo, sector_id, codigo_barras):
    if not codigo_barras or str(codigo_barras).strip() == "":
        codigo_barras = None
    margen = obtener_margen_sector(sector_id)
    precio = round((costo or 0.0) + ((costo or 0.0) * (margen or 0.0)), 2)
    

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO productos (codigo, nombre, cantidad, costo, sector_id, precio, codigo_barras, movimientos)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (codigo, nombre, cantidad, costo, sector_id, precio, codigo_barras, 0))
    producto_id = cur.lastrowid
    conn.commit()
    conn.close()

    agregar_movimiento(producto_id, "INGRESO", cantidad, precio, detalles="Alta/Ingreso inicial")
    return producto_id

def agregar_o_actualizar_producto(codigo, nombre, cantidad, costo, sector_id=None, codigo_barras=""):
    """
    Upsert de producto:
      - Si existe un producto con el mismo 'codigo' → actualiza (suma 'cantidad', actualiza costo/sector/precio/barcode).
      - Si no existe → crea uno nuevo.
    Notas:
      - 'codigo_barras' vacío se normaliza a None para no chocar con el índice único condicional (IS NOT NULL).
      - 'precio' se recalcula con el margen del sector.
    """
    # Normalizar barcode vacío a NULL (permite múltiples productos sin código de barras)
    if not codigo_barras or str(codigo_barras).strip() == "":
        codigo_barras = None

    conn = get_connection()
    cur = conn.cursor()

    # Buscar por 'codigo' interno (clave de upsert)
    cur.execute("SELECT id, cantidad, costo, sector_id FROM productos WHERE codigo = ?", (codigo,))
    row = cur.fetchone()

    if row:
        # Existe → actualizar sumando stock y recalculando precio según costo/sector resultantes
        producto_id = row[0]
        current_cant = row[1] or 0
        current_costo = row[2] if row[2] is not None else 0.0
        current_sector = row[3]

        new_cant = current_cant + (cantidad or 0)

        # Mantener costo/sector existentes si no se pasan nuevos
        costo_final = costo if (costo is not None) else current_costo
        sector_final = sector_id if (sector_id is not None) else current_sector

        margen = obtener_margen_sector(sector_final)
        # Si no hay costo, precio=0.0 (evita None)
        precio = round((costo_final or 0.0) * (1 + (margen or 0.0)), 2) if costo_final is not None else 0.0

        cur.execute("""
            UPDATE productos
               SET cantidad = ?,
                   costo = ?,
                   sector_id = ?,
                   precio = ?,
                   codigo_barras = ?
             WHERE id = ?
        """, (new_cant, costo_final, sector_final, precio, codigo_barras, producto_id))
        conn.commit()
        conn.close()

        # Registrar movimiento de ingreso por alta/import
        agregar_movimiento(producto_id, "INGRESO", cantidad or 0, precio, detalles="Ingreso por import/alta")
        return producto_id
    else:
        # No existe → crear nuevo (agregar_producto también normaliza y registra movimiento)
        conn.close()
        return agregar_producto(codigo, nombre, cantidad or 0, costo, sector_id, codigo_barras)

def eliminar_producto(id_producto):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT nombre, precio FROM productos WHERE id=?", (id_producto,))
    row = cur.fetchone()
    cur.execute("DELETE FROM productos WHERE id=?", (id_producto,))
    conn.commit()
    conn.close()

    if row:
         agregar_movimiento(None, "ELIM", 0, row[1] or 0, detalles=f"Eliminado: {row[0]}")

def obtener_productos():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT p.id, p.codigo, p.nombre, p.cantidad, p.costo,
               COALESCE(s.nombre, '') as sector, p.precio, COALESCE(p.codigo_barras, ''), p.movimientos
        FROM productos p
        LEFT JOIN sectores s ON p.sector_id = s.id
        ORDER BY p.nombre COLLATE NOCASE
    """)
    data = cur.fetchall()
    conn.close()
    return data

# --- NUEVO: búsqueda flexible de productos (prioriza match exacto por código / barcode)
def buscar_productos(query, limit=50):
    """
    Devuelve una lista de productos que coincidan con:
    - código exacto o código de barras exacto (flujo scanner)
    - o NOMBRE/CÓDIGO/BARCODE que CONTENGAN el texto (búsqueda manual)
    Columnas devueltas = igual que obtener_productos()
    """
    q = (query or "").strip()
    if not q:
        return []

    conn = get_connection()
    cur = conn.cursor()

    # 1) Match exacto por código interno o código de barras
    cur.execute("""
        SELECT p.id, p.codigo, p.nombre, p.cantidad, p.costo,
               COALESCE(s.nombre, '') as sector, p.precio, COALESCE(p.codigo_barras, ''), p.movimientos
        FROM productos p
        LEFT JOIN sectores s ON p.sector_id = s.id
        WHERE LOWER(p.codigo) = LOWER(?)
           OR LOWER(COALESCE(p.codigo_barras,'')) = LOWER(?)
        ORDER BY p.nombre COLLATE NOCASE
        LIMIT ?
    """, (q, q, limit))
    exactos = cur.fetchall()
    if exactos:
        conn.close()
        return exactos

    # 2) Búsqueda por contiene (insensible a mayúsculas)
    like = f"%{q}%"
    cur.execute("""
        SELECT p.id, p.codigo, p.nombre, p.cantidad, p.costo,
               COALESCE(s.nombre, '') as sector, p.precio, COALESCE(p.codigo_barras, ''), p.movimientos
        FROM productos p
        LEFT JOIN sectores s ON p.sector_id = s.id
        WHERE p.nombre LIKE ? COLLATE NOCASE
           OR p.codigo LIKE ? COLLATE NOCASE
           OR COALESCE(p.codigo_barras,'') LIKE ? COLLATE NOCASE
        ORDER BY p.nombre COLLATE NOCASE
        LIMIT ?
    """, (like, like, like, limit))
    data = cur.fetchall()
    conn.close()
    return data

# -----------------------------
# STOCK / VENTAS
# -----------------------------
def modificar_stock(producto_id, cantidad_cambio, detalles=""):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT cantidad, precio FROM productos WHERE id=?", (producto_id,))
    prod = cur.fetchone()
    if not prod:
        conn.close()
        return False

    nueva = (prod[0] or 0) + cantidad_cambio
    if nueva < 0:
        conn.close()
        return False

    cur.execute("UPDATE productos SET cantidad=? WHERE id=?", (nueva, producto_id))
    conn.commit()
    conn.close()

    tipo = "VENTA" if cantidad_cambio < 0 else "INGRESO"
    agregar_movimiento(producto_id, tipo, cantidad_cambio, prod[1] or 0.0, detalles=detalles)
    return True

# -----------------------------
# CLIENTES (nueva funcionalidad FASE 2)
# -----------------------------
def agregar_cliente(nombre, telefono="", direccion="", notas=""):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("INSERT INTO clientes (nombre, telefono, direccion, notas) VALUES (?, ?, ?, ?)",
                (nombre, telefono, direccion, notas))
    cid = cur.lastrowid
    conn.commit()
    conn.close()
    return cid

def editar_cliente(cid, nombre, telefono="", direccion="", notas=""):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE clientes SET nombre=?, telefono=?, direccion=?, notas=? WHERE id=?",
                (nombre, telefono, direccion, notas, cid))
    conn.commit()
    conn.close()

def eliminar_cliente(cid):
    conn = get_connection()
    cur = conn.cursor()
    # OJO: si hay ventas asociadas, dejar la referencia o prevenir eliminación según política
    cur.execute("DELETE FROM clientes WHERE id=?", (cid,))
    conn.commit()
    conn.close()

def obtener_clientes():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, nombre, telefono, direccion, notas FROM clientes ORDER BY nombre")
    data = cur.fetchall()
    conn.close()
    return data

def obtener_clientes_con_saldo():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT c.id, c.nombre, COALESCE(SUM(v.total),0) as deuda, COUNT(v.id) as cant_pendientes
        FROM clientes c
        LEFT JOIN ventas v ON v.cliente_id = c.id AND v.estado='PENDIENTE'
        GROUP BY c.id
        ORDER BY deuda DESC
    """)
    data = cur.fetchall()
    conn.close()
    return data

# -----------------------------
# VENTAS
# -----------------------------
def registrar_venta(items, tipo_pago, cliente=None, efectivo_recibido=None):
    """
    items: list of dicts: {"producto_id": id, "cantidad": n, "precio_unitario": p}
    tipo_pago: 'Efectivo'|'Transferencia'|'QR'|'Pendiente'
    cliente: None | cliente_id (int) | cliente_nombre (str)
    """
    conn = get_connection()
    cur = conn.cursor()
    try:
        total = round(sum(it["cantidad"] * it["precio_unitario"] for it in items), 2)
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        estado = "PAGADO" if tipo_pago != "Pendiente" else "PENDIENTE"
        vuelto = None

        # resolver cliente: obtener cliente_id y cliente_text
        cliente_id = None
        cliente_text = ""
        if cliente is not None:
            if isinstance(cliente, int):
                cliente_id = cliente
                cur.execute("SELECT nombre FROM clientes WHERE id=?", (cliente_id,))
                r = cur.fetchone()
                cliente_text = r[0] if r else ""
            else:
                # string
                cliente_text = str(cliente)

        # Iniciar transacción
        cur.execute("BEGIN")
        cur.execute("""
            INSERT INTO ventas (fecha, tipo_pago, estado, total, efectivo_recibido, vuelto, cliente, cliente_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (fecha, tipo_pago, estado, total, efectivo_recibido, vuelto, cliente_text, cliente_id))
        venta_id = cur.lastrowid

        # procesar items: verificar stock y descontar
        for it in items:
            pid = it["producto_id"]
            cant = it["cantidad"]
            precio_unit = it["precio_unitario"]
            subtotal = round(cant * precio_unit, 2)

            # verificar stock actual
            cur.execute("SELECT cantidad, precio FROM productos WHERE id=?", (pid,))
            prod = cur.fetchone()
            if not prod:
                raise Exception(f"Producto id {pid} no encontrado")
            current = prod[0] or 0
            if current - cant < 0:
                raise Exception(f"Stock insuficiente para {pid}")

            # descontar stock
            cur.execute("UPDATE productos SET cantidad=? WHERE id=?", (current - cant, pid))

            # insertar item
            cur.execute("""
                INSERT INTO venta_items (venta_id, producto_id, cantidad, precio_unitario, subtotal)
                VALUES (?, ?, ?, ?, ?)
            """, (venta_id, pid, cant, precio_unit, subtotal))

            # registrar movimiento tipo VENTA (cantidad negativa)
            cur.execute("INSERT INTO movimientos (producto_id, tipo, cambio, precio_unitario, fecha, detalles) VALUES (?, ?, ?, ?, ?, ?)",
                        (pid, "VENTA", -cant, precio_unit, fecha, f"Venta ID {venta_id}"))

        conn.commit()
        conn.close()
        return True, venta_id
    except Exception as e:
        conn.rollback()
        conn.close()
        return False, str(e)

def obtener_ventas(fecha_inicio=None, fecha_fin=None, estado=None):
    conn = get_connection()
    cur = conn.cursor()
    q = """
        SELECT v.id, v.fecha, v.tipo_pago, v.estado, v.total, v.efectivo_recibido, v.vuelto,
               COALESCE(c.nombre, v.cliente) as cliente
        FROM ventas v
        LEFT JOIN clientes c ON v.cliente_id = c.id
        WHERE 1=1
    """
    params = []
    if fecha_inicio and fecha_fin:
        q += " AND date(v.fecha) BETWEEN date(?) AND date(?)"
        params.extend([fecha_inicio, fecha_fin])
    if estado:
        q += " AND v.estado = ?"
        params.append(estado)
    q += " ORDER BY v.fecha DESC"
    cur.execute(q, tuple(params))
    data = cur.fetchall()
    conn.close()
    return data

def obtener_items_venta(venta_id):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT vi.id, vi.producto_id, p.nombre, vi.cantidad, vi.precio_unitario, vi.subtotal
        FROM venta_items vi
        JOIN productos p ON p.id = vi.producto_id
        WHERE vi.venta_id = ?
    """, (venta_id,))
    data = cur.fetchall()
    conn.close()
    return data

def obtener_ventas_pendientes():
    return obtener_ventas(estado="PENDIENTE")

def marcar_venta_pagada(venta_id, tipo_pago_nuevo=None, recibido=None):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT estado, total, cliente_id FROM ventas WHERE id=?", (venta_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return False, "Venta no encontrada"
    estado_actual = row[0]
    total = row[1] or 0.0
    cliente_id = row[2]

    if estado_actual == "PAGADO":
        conn.close()
        return False, "Venta ya está pagada"

    nuevo_tipo = tipo_pago_nuevo if tipo_pago_nuevo else "Efectivo"
    vuelto = None
    if recibido is not None:
        vuelto = round(recibido - total, 2)

    cur.execute("UPDATE ventas SET estado=?, tipo_pago=?, efectivo_recibido=?, vuelto=? WHERE id=?",
                ("PAGADO", nuevo_tipo, recibido, vuelto, venta_id))

    # insertar registro de cobro
    cur.execute("INSERT INTO cobros (venta_id, cliente_id, fecha, monto, tipo_pago) VALUES (?, ?, ?, ?, ?)",
                (venta_id, cliente_id, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), total, nuevo_tipo))

    conn.commit()
    conn.close()
    return True, "Venta marcada como pagada"

def reembolsar_venta(venta_id, items_to_refund=None):
    conn = get_connection()
    cur = conn.cursor()
    try:
        # Buscar la venta
        cur.execute("SELECT id, fecha, total FROM ventas WHERE id=?", (venta_id,))
        venta = cur.fetchone()
        if not venta:
            conn.close()
            return False, "Venta no encontrada"

        # Traer los ítems de la venta
        if items_to_refund is None:
            cur.execute("""
                SELECT id, producto_id, cantidad, precio_unitario, subtotal
                FROM venta_items WHERE venta_id=?
            """, (venta_id,))
            items = cur.fetchall()
        else:
            placeholders = ",".join("?" for _ in items_to_refund)
            cur.execute(f"""
                SELECT id, producto_id, cantidad, precio_unitario, subtotal
                FROM venta_items
                WHERE id IN ({placeholders})
            """, tuple(items_to_refund))
            items = cur.fetchall()

        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        total_devuelto = 0.0

        for itm in items:
            vi_id, pid, cant, precio_unit, subtotal = itm
            total_devuelto += subtotal

            # Reponer stock
            cur.execute("UPDATE productos SET cantidad = cantidad + ? WHERE id=?", (cant, pid))

            # Registrar movimiento
            cur.execute("""
                INSERT INTO movimientos (producto_id, tipo, cambio, precio_unitario, fecha, detalles)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (pid, "REEMBOLSO", cant, precio_unit, fecha, f"Reembolso de venta {venta_id}"))

            # Eliminar el ítem de la venta
            cur.execute("DELETE FROM venta_items WHERE id=?", (vi_id,))

        # Actualizar el total de la venta
        cur.execute("UPDATE ventas SET total = total - ? WHERE id=?", (total_devuelto, venta_id))

        conn.commit()
        conn.close()
        return True, f"Reembolso procesado. Total devuelto: ${total_devuelto:,.2f}"
    except Exception as e:
        conn.rollback()
        conn.close()
        return False, str(e)



# -----------------------------
# Reportes: ventas agrupadas por tipo de pago
# -----------------------------
def ventas_resumen_por_tipo(fecha_inicio=None, fecha_fin=None):
    conn = get_connection()
    cur = conn.cursor()
    q = "SELECT tipo_pago, SUM(total) as total, COUNT(*) as cantidad FROM ventas WHERE 1=1"
    params = []
    if fecha_inicio and fecha_fin:
        q += " AND date(fecha) BETWEEN date(?) AND date(?)"
        params.extend([fecha_inicio, fecha_fin])
    q += " GROUP BY tipo_pago"
    cur.execute(q, tuple(params))
    data = cur.fetchall()
    conn.close()
    return data

def obtener_ventas_con_detalles():
    conn = get_connection()
    cur = conn.cursor()

    # Traer todas las ventas
    cur.execute("SELECT id, fecha, tipo_pago, estado, total FROM ventas ORDER BY fecha DESC")
    ventas = cur.fetchall()

    ventas_lista = []
    for v in ventas:
        venta = {
            "id": v[0],
            "fecha": v[1],
            "tipo_pago": v[2],
            "estado": v[3],
            "total": float(v[4]) if v[4] else 0.0,
        }

        # Traer los items de cada venta con ID real del item
        cur.execute("""
            SELECT vi.id, vi.producto_id, p.nombre, vi.cantidad, vi.precio_unitario, vi.subtotal
            FROM venta_items vi
            JOIN productos p ON vi.producto_id = p.id
            WHERE vi.venta_id=?
        """, (venta["id"],))
        items = cur.fetchall()

        venta["items"] = [
            {
                "id": it[0],             # ID real de venta_items
                "producto_id": it[1],
                "nombre": it[2],
                "cantidad": it[3],
                "precio": float(it[4]) if it[4] else 0.0,
                "subtotal": float(it[5]) if it[5] else 0.0,
            }
            for it in items
        ]

        ventas_lista.append(venta)

    conn.close()
    return ventas_lista


def obtener_producto_por_barcode(codigo_barras):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT p.id, p.codigo, p.nombre, p.cantidad, p.costo,
               COALESCE(s.nombre, '') as sector, p.precio, COALESCE(p.codigo_barras, ''), p.movimientos
        FROM productos p
        LEFT JOIN sectores s ON p.sector_id = s.id
        WHERE p.codigo_barras = ?
    """, (codigo_barras,))
    prod = cur.fetchone()
    conn.close()
    return prod

# -----------------------------
# GASTOS (Almacén y Personales)
# -----------------------------
def agregar_gasto(categoria, monto, detalle="", tipo="almacen"):
    conn = get_connection()
    cur = conn.cursor()
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur.execute("""
        INSERT INTO gastos (fecha, categoria, monto, detalle, tipo)
        VALUES (?, ?, ?, ?, ?)
    """, (fecha, categoria, monto, detalle, tipo))
    conn.commit()
    conn.close()

def obtener_gastos(tipo="almacen", fecha_inicio=None, fecha_fin=None):
    conn = get_connection()
    cur = conn.cursor()
    query = "SELECT id, fecha, categoria, monto, detalle FROM gastos WHERE tipo=?"
    params = [tipo]

    if fecha_inicio and fecha_fin:
        query += " AND date(fecha) BETWEEN date(?) AND date(?)"
        params.extend([fecha_inicio, fecha_fin])
    elif fecha_inicio:
        query += " AND date(fecha) >= date(?)"
        params.append(fecha_inicio)
    elif fecha_fin:
        query += " AND date(fecha) <= date(?)"
        params.append(fecha_fin)

    query += " ORDER BY fecha DESC"
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    conn.close()
    return rows

def obtener_resumen_gastos(tipo="almacen", fecha_inicio=None, fecha_fin=None):
    conn = get_connection()
    cur = conn.cursor()
    query = "SELECT categoria, SUM(monto) FROM gastos WHERE tipo=?"
    params = [tipo]

    if fecha_inicio and fecha_fin:
        query += " AND date(fecha) BETWEEN date(?) AND date(?)"
        params.extend([fecha_inicio, fecha_fin])
    elif fecha_inicio:
        query += " AND date(fecha) >= date(?)"
        params.append(fecha_inicio)
    elif fecha_fin:
        query += " AND date(fecha) <= date(?)"
        params.append(fecha_fin)

    query += " GROUP BY categoria ORDER BY SUM(monto) DESC"
    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    conn.close()
    return rows

def eliminar_gasto(gasto_id):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM gastos WHERE id=?", (gasto_id,))
    conn.commit()
    conn.close()


def exportar_gastos_excel(tipo="almacen", filename="gastos.xlsx", fecha_inicio=None, fecha_fin=None):
    gastos = obtener_gastos(tipo, fecha_inicio, fecha_fin)
    resumen = obtener_resumen_gastos(tipo, fecha_inicio, fecha_fin)

    wb = Workbook()

    # === Hoja Detalle ===
    ws1 = wb.active
    ws1.title = "Detalle"

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Encabezado
    ws1.merge_cells("A1:E1")
    rango_txt = ""
    if fecha_inicio and fecha_fin:
        rango_txt = f" ({fecha_inicio} a {fecha_fin})"
    elif fecha_inicio:
        rango_txt = f" (desde {fecha_inicio})"
    elif fecha_fin:
        rango_txt = f" (hasta {fecha_fin})"
    ws1["A1"] = f"Reporte de gastos - {tipo.capitalize()}{rango_txt}"
    ws1["A1"].font = Font(size=16, bold=True, color="1F497D")
    ws1["A1"].alignment = Alignment(horizontal="center")

    # Encabezados de tabla
    headers = ["ID", "Fecha", "Categoría", "Monto", "Detalle"]
    ws1.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row=2, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws1.column_dimensions[get_column_letter(col)].width = 20

    # Datos
    for g in gastos:
        ws1.append(g)

    # Formato y bordes
    for row in ws1.iter_rows(min_row=3, max_row=ws1.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            if cell.column == 4:  # monto
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # Total General en Detalle
    total_detalle = sum(g[3] for g in gastos) if gastos else 0
    total_row = ws1.max_row + 2
    ws1["C{}".format(total_row)] = "TOTAL GENERAL"
    ws1["C{}".format(total_row)].font = Font(bold=True, color="C00000")
    ws1["D{}".format(total_row)] = total_detalle
    ws1["D{}".format(total_row)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws1["D{}".format(total_row)].font = Font(bold=True, color="C00000")

    # === Hoja Resumen ===
    ws2 = wb.create_sheet(title="Resumen")

    ws2.merge_cells("A1:C1")
    ws2["A1"] = f"Resumen de gastos - {tipo.capitalize()}{rango_txt}"
    ws2["A1"].font = Font(size=16, bold=True, color="1F497D")
    ws2["A1"].alignment = Alignment(horizontal="center")

    ws2.append(["Categoría", "Total"])
    for col in range(1, 3):
        cell = ws2.cell(row=2, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(col)].width = 25

    total_general = 0
    for r in resumen:
        ws2.append(r)
        total_general += r[1]

    for row in ws2.iter_rows(min_row=3, max_row=ws2.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            if cell.column == 2:
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # Total general en Resumen
    ws2["A{}".format(ws2.max_row + 2)] = "TOTAL GENERAL"
    ws2["A{}".format(ws2.max_row)].font = Font(bold=True, color="C00000")
    ws2["B{}".format(ws2.max_row)] = total_general
    ws2["B{}".format(ws2.max_row)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws2["B{}".format(ws2.max_row)].font = Font(bold=True, color="C00000")

    # Gráfico circular
    if ws2.max_row > 3:
        pie = PieChart()
        pie.title = f"Gastos por categoría ({tipo})"
        labels = Reference(ws2, min_col=1, min_row=3, max_row=ws2.max_row - 2)
        data = Reference(ws2, min_col=2, min_row=2, max_row=ws2.max_row - 2)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 12
        pie.width = 12

        # Crear etiquetas de datos si no existen
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCategory = True

        ws2.add_chart(pie, "D4")

    wb.save(filename)
    return filename

# --------- Tickets ---------

# Elegí el formato por defecto: "termico" (58 mm) o "a4"
FORMATO_TICKET = "termico"

def _datos_venta_y_items(venta_id):
    """Devuelve (venta, items) con nombre de cliente ya resuelto."""
    conn = get_connection()
    cur = conn.cursor()
    # Traemos la venta y el nombre del cliente (si existe)
    cur.execute("""
        SELECT v.fecha, v.tipo_pago, v.total, v.efectivo_recibido, v.vuelto,
               COALESCE(c.nombre, 'Consumidor Final') AS cliente_nombre
        FROM ventas v
        LEFT JOIN clientes c ON c.id = v.cliente_id
        WHERE v.id = ?
    """, (venta_id,))
    venta = cur.fetchone()
    if not venta:
        conn.close()
        return None, None

    # Traemos items con nombre de producto
    cur.execute("""
        SELECT p.nombre, vi.cantidad, vi.precio_unitario, vi.subtotal
        FROM venta_items vi
        JOIN productos p ON p.id = vi.producto_id
        WHERE vi.venta_id = ?
    """, (venta_id,))
    items = cur.fetchall()
    conn.close()
    return venta, items

def generar_ticket_a4(venta_id, ruta):
    venta, items = _datos_venta_y_items(venta_id)
    if not venta:
        return None

    ancho, alto = A4
    c = canvas.Canvas(ruta, pagesize=A4)
    y = alto - 20*mm

    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(ancho/2, y, "Mi Almacén")
    y -= 10*mm

    c.setFont("Helvetica", 10)
    c.drawString(20*mm, y, f"Fecha: {venta[0]}")
    y -= 5*mm
    c.drawString(20*mm, y, f"Cliente: {venta[5]}")
    y -= 10*mm

    c.drawString(20*mm, y, "Cant  Producto              P.Unit   Subtotal")
    y -= 5*mm
    for nombre, cant, precio, subtotal in items:
        c.drawString(20*mm, y, f"{cant:>3}  {nombre[:20]:<20} {precio:>7.2f}  {subtotal:>7.2f}")
        y -= 5*mm

    y -= 10*mm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(20*mm, y, f"TOTAL: ${venta[2]:.2f}")
    y -= 7*mm
    c.setFont("Helvetica", 10)
    c.drawString(20*mm, y, f"Pago: {venta[1]}")
    if venta[1] == "Efectivo":
        y -= 5*mm
        c.drawString(20*mm, y, f"Recibido: ${venta[3] or 0:.2f}  Vuelto: ${venta[4] or 0:.2f}")

    c.save()
    return ruta

def generar_ticket_termico(venta_id, ruta):
    """Ticket térmico 58 mm de ancho."""
    venta, items = _datos_venta_y_items(venta_id)
    if not venta:
        return None

    ancho = 58 * mm
    alto = 200 * mm  # suficiente para ~30 líneas
    c = canvas.Canvas(ruta, pagesize=(ancho, alto))
    y = alto - 10 * mm

    # Encabezado
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(ancho/2, y, "Mi Almacén")
    y -= 6 * mm

    c.setFont("Helvetica", 7)
    c.drawString(4 * mm, y, f"Fecha: {venta[0]}")
    y -= 4 * mm
    c.drawString(4 * mm, y, f"Cliente: {venta[5]}")
    y -= 6 * mm

    # Productos
    c.setFont("Helvetica-Bold", 7)
    c.drawString(4 * mm, y, "Cant Prod         Subtot")
    y -= 4 * mm
    c.setFont("Helvetica", 7)

    for nombre, cant, precio, subtotal in items:
        nombre_corto = (nombre[:10] + "...") if len(nombre) > 12 else nombre
        c.drawString(4 * mm, y, f"{cant:>2}  {nombre_corto:<12} {subtotal:>6.2f}")
        y -= 4 * mm

    # Total y pago
    y -= 4 * mm
    c.setFont("Helvetica-Bold", 8)
    c.drawString(4 * mm, y, f"TOTAL: ${venta[2]:.2f}")
    y -= 5 * mm

    c.setFont("Helvetica", 7)
    c.drawString(4 * mm, y, f"Pago: {venta[1]}")
    if venta[1] == "Efectivo":
        y -= 4 * mm
        c.drawString(4 * mm, y, f"Recibido: ${venta[3] or 0:.2f}")
        y -= 4 * mm
        c.drawString(4 * mm, y, f"Vuelto:   ${venta[4] or 0:.2f}")

    # Mensaje final
    y -= 8 * mm
    c.setFont("Helvetica-Oblique", 7)
    c.drawCentredString(ancho/2, y, "¡Gracias por su compra!")

    c.save()
    return ruta

def generar_ticket(venta_id, formato=None):
    """Wrapper: crea carpeta, arma ruta y delega según formato."""
    os.makedirs("tickets", exist_ok=True)
    ruta = os.path.join("tickets", f"ticket_{venta_id}.pdf")
    fmt = (formato or FORMATO_TICKET).lower()

    if fmt == "a4":
        return generar_ticket_a4(venta_id, ruta)
    else:
        return generar_ticket_termico(venta_id, ruta)

def guardar_carrito_temporal(lista_items):
    """Guarda el carrito actual en la tabla temporal."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM carrito_temporal")
    for it in lista_items:
        cursor.execute("""
            INSERT INTO carrito_temporal (producto_id, codigo, nombre, cantidad, precio_unitario)
            VALUES (?, ?, ?, ?, ?)
        """, (it["producto_id"], it["codigo"], it["nombre"], it["cantidad"], it["precio_unitario"]))
    conn.commit()
    conn.close()

def obtener_carrito_temporal():
    """Devuelve lista de items del carrito temporal."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT producto_id, codigo, nombre, cantidad, precio_unitario FROM carrito_temporal")
    rows = cursor.fetchall()
    conn.close()
    return [
        {"producto_id": r[0], "codigo": r[1], "nombre": r[2], "cantidad": r[3], "precio_unitario": r[4]}
        for r in rows
    ]

def limpiar_carrito_temporal():
    """Borra el carrito temporal de la base."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM carrito_temporal")
    conn.commit()
    conn.close()
    
# -----------------------------
# CATEGORÍAS DE GASTOS
# -----------------------------
def agregar_categoria_gasto(nombre, tipo="almacen"):
    """Agrega una categoría de gasto si no existe."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS categorias_gasto (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT UNIQUE NOT NULL,
            tipo TEXT NOT NULL CHECK(tipo IN ('almacen','personal'))
        )
    """)
    cur.execute("INSERT OR IGNORE INTO categorias_gasto (nombre, tipo) VALUES (?, ?)", (nombre, tipo))
    conn.commit()
    conn.close()

def obtener_categorias_gasto(tipo="almacen"):
    """Devuelve lista de nombres de categorías según tipo."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT nombre FROM categorias_gasto WHERE tipo=? ORDER BY nombre", (tipo,))
    rows = [r[0] for r in cur.fetchall()]
    conn.close()
    return rows


def _hash_password(plain: str, iterations: int = 200_000) -> str:
    if plain is None:
        plain = ""
    salt = _os.urandom(16)
    dk = _hashlib.pbkdf2_hmac("sha256", plain.encode("utf-8"), salt, iterations)
    return "pbkdf2_sha256$%d$%s$%s" % (
        iterations,
        _binascii.hexlify(salt).decode(),
        _binascii.hexlify(dk).decode()
    )

# Reemplazar toda tu función _verify_password por esta
import binascii as _binascii, base64 as _base64, secrets as _secrets, hashlib as _hashlib

def _verify_password(plain: str, stored: str) -> bool:
    """
    Verifica contraseñas en formato:
      pbkdf2_sha256$<iter>$<salt>$<hash>
    Admitiendo variantes:
      - Separadores ":" o "$"
      - Salt/hash en hex o base64
      - Espacios accidentales
    También acepta texto plano legado (igualdad directa).
    """
    if not stored:
        return False

    # Compatibilidad texto plano
    if not stored.startswith("pbkdf2_sha256"):
        return plain == stored

    try:
        s = stored.strip()

        # Normalizar separador a "$"
        s = s.replace(":", "$")

        parts = s.split("$")
        # Esperamos: ["pbkdf2_sha256", "<iter>", "<salt>", "<hash>"]
        if len(parts) < 4:
            return False

        _, it_str, salt_s, hash_s = parts[0], parts[1], parts[2], parts[3]

        # Decode helper: primero intento hex, si falla intento base64
        def _decode(maybe_hex_or_b64: str):
            try:
                return _binascii.unhexlify(maybe_hex_or_b64.encode())
            except Exception:
                try:
                    return _base64.b64decode(maybe_hex_or_b64.encode(), validate=True)
                except Exception:
                    return None

        salt = _decode(salt_s)
        expected = _decode(hash_s)
        if salt is None or expected is None:
            return False

        iterations = int(it_str)
        dk = _hashlib.pbkdf2_hmac("sha256", (plain or "").encode("utf-8"), salt, iterations)

        # compare_digest a prueba de timing
        return _secrets.compare_digest(dk, expected)
    except Exception:
        return False
    
def verificar_usuario(usuario: str, password: str):
    usuario = (usuario or "").strip()
    password = (password or "")
    conn = get_connection()
    cur = conn.cursor()
    # ⚠️ Selecciona SOLO por usuario; la contraseña se valida en Python
    cur.execute("SELECT password, rol FROM usuarios WHERE usuario=?", (usuario,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return None
    stored, rol = row
    return rol if _verify_password(password, stored) else None


def crear_usuario(usuario, password, rol="user"):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("INSERT INTO usuarios (usuario, password, rol) VALUES (?, ?, ?)",
                (usuario, _hash_password(password), rol))
    conn.commit()
    conn.close()

# -----------------------------
# USUARIOS
# -----------------------------
def obtener_usuarios():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, usuario, rol FROM usuarios ORDER BY usuario")
    data = cur.fetchall()
    conn.close()
    return data

def agregar_usuario(usuario, password, rol):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("INSERT INTO usuarios (usuario, password, rol) VALUES (?, ?, ?)",
                (usuario, _hash_password(password), rol))
    conn.commit()
    conn.close()

def editar_usuario(uid, usuario, password, rol):
    conn = get_connection()
    cur = conn.cursor()
    if password:
        cur.execute("UPDATE usuarios SET usuario=?, password=?, rol=? WHERE id=?",
                    (usuario, _hash_password(password), rol, uid))
    else:
        cur.execute("UPDATE usuarios SET usuario=?, rol=? WHERE id=?",
                    (usuario, rol, uid))
    conn.commit()
    conn.close()

def eliminar_usuario(uid):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM usuarios WHERE id=?", (uid,))
    conn.commit()
    conn.close()